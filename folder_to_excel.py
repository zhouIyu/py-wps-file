#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
文件夹图片到Excel工具
将文件夹中的二级目录名称作为sheet，并将目录内的图片插入到Excel文档中
"""

import os
import sys
import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing import image as openpyxl_image
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from PIL import Image
import logging
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading

# 支持的图片格式
SUPPORTED_IMAGE_FORMATS = {'.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp'}

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def is_image_file(file_path):
    """检查文件是否为支持的图片格式"""
    return file_path.suffix.lower() in SUPPORTED_IMAGE_FORMATS


def resize_image_for_excel(image_path, max_width=400, max_height=300):
    """
    调整图片尺寸以适合Excel单元格
    
    Args:
        image_path (Path): 图片文件路径
        max_width (int): 最大宽度（像素）
        max_height (int): 最大高度（像素）
    
    Returns:
        str: 临时调整后的图片路径
    """
    try:
        with Image.open(image_path) as img:
            # 计算缩放比例
            width_ratio = max_width / img.width if img.width > max_width else 1
            height_ratio = max_height / img.height if img.height > max_height else 1
            ratio = min(width_ratio, height_ratio)
            
            if ratio < 1:
                new_size = (int(img.width * ratio), int(img.height * ratio))
                img_resized = img.resize(new_size, Image.Resampling.LANCZOS)
                
                # 保存临时文件
                temp_path = image_path.parent / f"temp_{image_path.stem}.png"
                img_resized.save(temp_path, "PNG")
                return str(temp_path)
            else:
                return str(image_path)
                
    except Exception as e:
        logger.warning(f"调整图片尺寸失败 {image_path}: {e}")
        return str(image_path)


def create_excel_document(folder_path, output_path="图片表格.xlsx", progress_callback=None):
    """
    创建Excel文档，将文件夹中的二级目录作为sheet，图片插入到对应sheet中
    
    Args:
        folder_path (str): 源文件夹路径
        output_path (str): 输出Excel文档路径
        progress_callback (function): 进度回调函数
    """
    folder_path = Path(folder_path)
    
    if not folder_path.exists():
        logger.error(f"文件夹不存在: {folder_path}")
        return False
    
    if not folder_path.is_dir():
        logger.error(f"路径不是文件夹: {folder_path}")
        return False
    
    # 创建Excel工作簿
    wb = Workbook()
    # 删除默认的Sheet
    if wb.active:
        wb.remove(wb.active)
    
    # 统计信息
    total_directories = 0
    total_images = 0
    temp_files = []  # 记录临时文件，用于清理
    
    # 获取所有二级目录
    subdirectories = []
    try:
        for item in folder_path.iterdir():
            if item.is_dir():
                subdirectories.append(item)
    except PermissionError:
        logger.error(f"没有权限访问文件夹: {folder_path}")
        return False
    
    # 按目录名排序
    subdirectories.sort(key=lambda x: x.name)
    
    if not subdirectories:
        logger.warning(f"在 {folder_path} 中没有找到子目录")
        # 创建一个空的sheet
        ws = wb.create_sheet("无内容")
        ws['A1'] = "没有找到子目录"
        wb.save(output_path)
        return True
    
    logger.info(f"找到 {len(subdirectories)} 个子目录")
    
    # 先扫描所有目录，收集有图片的目录信息
    valid_directories = []
    for subdir in subdirectories:
        # 获取目录中的所有图片文件
        image_files = []
        try:
            for file_path in subdir.iterdir():
                if file_path.is_file() and is_image_file(file_path):
                    image_files.append(file_path)
        except PermissionError:
            logger.warning(f"没有权限访问目录: {subdir}")
            continue
        
        if image_files:  # 只记录有图片的目录
            valid_directories.append((subdir, image_files))
    
    if not valid_directories:
        logger.warning(f"没有找到包含图片的目录")
        ws = wb.create_sheet("无内容")
        ws['A1'] = "没有找到包含图片的目录"
        wb.save(output_path)
        return True
    
    # 遍历每个有效目录
    for i, (subdir, image_files) in enumerate(valid_directories):
        logger.info(f"处理目录: {subdir.name}")
        
        # 清理sheet名称，Excel sheet名称不能包含特殊字符
        sheet_name = subdir.name.replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('[', '_').replace(']', '_')
        # 限制sheet名称长度
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        
        # 创建新的工作表
        ws = wb.create_sheet(sheet_name)
        
        # 按文件名排序图片
        image_files.sort(key=lambda x: x.name)
        
        total_directories += 1
        
        # 设置标题
        ws['A1'] = subdir.name
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # 合并标题单元格（只使用A列）
        ws.merge_cells('A1:A1')
        
        # 设置列宽
        ws.column_dimensions['A'].width = 40  # A列放图片，设置较宽
        
        logger.info(f"  找到 {len(image_files)} 张图片")
        
        # 添加图片 - 垂直排列在一列中
        current_row = 3  # 从第3行开始添加图片
        col = 1  # 固定在第1列（A列）
        row_spacing = 3  # 每张图片之间的行间距
        
        for j, img_path in enumerate(image_files):
            try:
                logger.info(f"    添加图片: {img_path.name}")
                
                # 调整图片尺寸
                temp_img_path = resize_image_for_excel(img_path)
                if temp_img_path != str(img_path):
                    temp_files.append(temp_img_path)
                
                # 插入图片
                img = openpyxl_image.Image(temp_img_path)
                
                # 设置图片位置（所有图片都在A列）
                cell_position = f"A{current_row}"
                ws.add_image(img, cell_position)
                
                # 设置行高以适应图片
                ws.row_dimensions[current_row].height = 200
                
                # 准备下一张图片的位置
                current_row += row_spacing
                
                total_images += 1
                
            except Exception as e:
                logger.error(f"    添加图片失败 {img_path.name}: {e}")
        
        # 更新进度
        if progress_callback:
            progress = (i + 1) / len(valid_directories) * 100
            progress_callback(progress)
    
    # 保存Excel文档
    try:
        wb.save(output_path)
        logger.info(f"Excel文档已保存: {Path(output_path).absolute()}")
        logger.info(f"统计信息 - Sheet数量: {total_directories}，图片: {total_images}")
        
        # 清理临时文件
        for temp_file in temp_files:
            try:
                os.remove(temp_file)
            except:
                pass
        
        return True
    except Exception as e:
        logger.error(f"保存Excel文档失败: {e}")
        return False


class FolderToExcelGUI:
    """GUI界面"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("文件夹图片转Excel工具")
        self.root.geometry("500x400")
        self.root.resizable(True, True)
        
        self.folder_path = tk.StringVar()
        self.output_path = tk.StringVar(value="图片表格.xlsx")
        
        self.setup_ui()
        
    def setup_ui(self):
        """设置UI界面"""
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky="nsew")
        
        # 选择文件夹
        ttk.Label(main_frame, text="选择包含图片的文件夹:").grid(row=0, column=0, sticky=tk.W, pady=(0, 5))
        
        folder_frame = ttk.Frame(main_frame)
        folder_frame.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 15))
        folder_frame.columnconfigure(0, weight=1)
        
        ttk.Entry(folder_frame, textvariable=self.folder_path, width=50).grid(row=0, column=0, sticky="ew", padx=(0, 5))
        ttk.Button(folder_frame, text="浏览", command=self.browse_folder).grid(row=0, column=1)
        
        # 输出文件
        ttk.Label(main_frame, text="输出Excel文件名:").grid(row=2, column=0, sticky=tk.W, pady=(0, 5))
        
        output_frame = ttk.Frame(main_frame)
        output_frame.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(0, 15))
        output_frame.columnconfigure(0, weight=1)
        
        ttk.Entry(output_frame, textvariable=self.output_path, width=50).grid(row=0, column=0, sticky="ew", padx=(0, 5))
        ttk.Button(output_frame, text="另存为", command=self.browse_output).grid(row=0, column=1)
        
        # 转换按钮
        self.convert_btn = ttk.Button(main_frame, text="开始转换", command=self.start_conversion)
        self.convert_btn.grid(row=4, column=0, columnspan=2, pady=(0, 15))
        
        # 进度条
        self.progress = ttk.Progressbar(main_frame, mode='determinate')
        self.progress.grid(row=5, column=0, columnspan=2, sticky="ew", pady=(0, 10))
        
        # 状态标签
        self.status_label = ttk.Label(main_frame, text="请选择文件夹开始转换")
        self.status_label.grid(row=6, column=0, columnspan=2, sticky=tk.W)
        
        # 日志文本框
        ttk.Label(main_frame, text="处理日志:").grid(row=7, column=0, sticky=tk.W, pady=(15, 5))
        
        log_frame = ttk.Frame(main_frame)
        log_frame.grid(row=8, column=0, columnspan=2, sticky="nsew", pady=(0, 10))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        self.log_text = tk.Text(log_frame, height=10, width=60)
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        
        self.log_text.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        
        # 配置权重
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(8, weight=1)
        
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        
    def browse_folder(self):
        """浏览文件夹"""
        folder = filedialog.askdirectory(title="选择包含图片的文件夹")
        if folder:
            self.folder_path.set(folder)
            
    def browse_output(self):
        """选择输出文件"""
        filename = filedialog.asksaveasfilename(
            title="保存Excel文件",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if filename:
            self.output_path.set(filename)
            
    def log_message(self, message):
        """添加日志消息"""
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
    def update_progress(self, value):
        """更新进度条"""
        self.progress['value'] = value
        self.status_label.config(text=f"处理中... {value:.1f}%")
        self.root.update_idletasks()
        
    def start_conversion(self):
        """开始转换"""
        if not self.folder_path.get():
            messagebox.showerror("错误", "请选择文件夹")
            return
            
        if not self.output_path.get():
            messagebox.showerror("错误", "请输入输出文件名")
            return
        
        # 禁用按钮
        self.convert_btn.config(state='disabled')
        self.progress['value'] = 0
        self.log_text.delete(1.0, tk.END)
        
        # 在新线程中执行转换
        thread = threading.Thread(target=self.convert_thread)
        thread.daemon = True
        thread.start()
        
    def convert_thread(self):
        """转换线程"""
        try:
            self.log_message(f"开始处理文件夹: {self.folder_path.get()}")
            
            success = create_excel_document(
                self.folder_path.get(),
                self.output_path.get(),
                self.update_progress
            )
            
            if success:
                self.log_message("转换完成!")
                self.status_label.config(text="转换完成!")
                messagebox.showinfo("成功", f"Excel文件已保存: {self.output_path.get()}")
            else:
                self.log_message("转换失败!")
                self.status_label.config(text="转换失败!")
                messagebox.showerror("错误", "转换失败，请查看日志")
                
        except Exception as e:
            self.log_message(f"错误: {str(e)}")
            self.status_label.config(text="转换失败!")
            messagebox.showerror("错误", f"转换失败: {str(e)}")
            
        finally:
            # 重新启用按钮
            self.convert_btn.config(state='normal')


def main():
    """主函数"""
    if len(sys.argv) > 1:
        # 命令行模式
        parser = argparse.ArgumentParser(description="将文件夹中的二级目录作为Excel sheet，图片插入到对应sheet中")
        parser.add_argument("folder_path", help="源文件夹路径")
        parser.add_argument("-o", "--output", default="图片表格.xlsx", 
                           help="输出Excel文档路径 (默认: 图片表格.xlsx)")
        
        args = parser.parse_args()
        
        # 检查文件夹是否存在
        if not os.path.exists(args.folder_path):
            print(f"错误: 文件夹不存在: {args.folder_path}")
            sys.exit(1)
        
        # 创建Excel文档
        success = create_excel_document(args.folder_path, args.output)
        
        if success:
            print(f"\n✓ 成功创建Excel文档: {args.output}")
        else:
            print("✗ 创建Excel文档失败")
            sys.exit(1)
    else:
        # GUI模式
        root = tk.Tk()
        app = FolderToExcelGUI(root)
        root.mainloop()


if __name__ == "__main__":
    main() 